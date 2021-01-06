from selenium import webdriver
import time
import os
import urllib
import requests
import shutil
import openpyxl
from selenium.webdriver.common.keys import Keys
from tqdm import tqdm

def write_excel_xls(path):
    wb=openpyxl.Workbook()
    wb.save(path)
    print("xlsx table created successfully ！") 
     
def write_excel_xls_append_norepeat(path, value):
    wb=openpyxl.load_workbook(path)
    ws = wb.active
    ws.append(value)
    rid = ws.max_row
    print("Successfully insert data number %d" % rid)
    wb.save(path)

#模拟浏览器向下滚的操作，来加载出更多内容
def Transfer_Clicks(browser):
    try:
        browser.execute_script("window.scrollBy(0,document.body.scrollHeight)", "")
    except:
        pass
    return "Transfer successfully \n"


def loginWeibo(username,password):
    driver.get('https://weibo.com')
    time.sleep(10)

    driver.find_element_by_id("loginname").send_keys(username)
    driver.find_element_by_xpath("//input[@type='password']").send_keys(password)
    driver.find_element_by_xpath("//div[@class='info_list login_btn']").click()


#下载图片
def img_download(pics,image_path,path):
    wb=openpyxl.load_workbook(path)
    ws=wb.active
    rows=ws.max_row
    file_path = image_path+"/"+str(rows)
    # print(file_path)
    # print(len(pics))
    for i in range(len(pics)):
        file_name = str(i)
        if (pics[i].get_attribute('src')).endswith('.jpg'):#保存jpg图片
             #是否有这个路径
            if not os.path.exists(file_path):
            #创建路径
                os.makedirs(file_path)
                #获得图片后缀
            file_suffix = os.path.splitext(pics[i].get_attribute('src'))[1]
            # print(file_suffix)
                #拼接图片名（包含路径）
            filename = '{}{}{}{}'.format(file_path,os.sep,file_name,file_suffix)
            # print(filename)
                #下载图片，并保存到文件夹中
            urllib.request.urlretrieve(pics[i].get_attribute('src'),filename=filename)
#插入数据
def insert_data(path,elems,userId):
    image_path=os.getcwd()+'/depression/csdn/img/'
    # print(len(elems))
    for elem in elems:
        # print(len(elems))
        weibo_content = elem.find_element_by_xpath(".//div[@class='WB_text W_f14' and @node-type='feed_list_content']").text
        # print(weibo_content)
        if weibo_content.find("转发") == -1:
            tmp = elem.find_elements_by_xpath(".//div[@class='WB_text W_f14' and @node-type='feed_list_content']//a[@class='WB_text_opt']")
            pics=elem.find_elements_by_xpath(".//div[@class='WB_media_wrap clearfix']//ul[@class='WB_media_a  WB_media_a_m1 clearfix' or @node-type='fl_pic_list']//li//img")
            
            if len(tmp)>0 and '展开全文' in tmp[0].text:
                tmp[0].send_keys(Keys.ENTER) 
                time.sleep(4)
                weibo_content = elem.find_element_by_xpath(".//div[@class='WB_text W_f14' and @node-type='feed_list_content_full']").text

                print('show full content')
            #消除标签
                tags = elem.find_elements_by_xpath(
                    ".//div[@class='WB_text W_f14' and @node-type='feed_list_content_full']/a")
            else:
                tags = elem.find_elements_by_xpath(
                    ".//div[@class='WB_text W_f14' and @node-type='feed_list_content']/a")
            # print(weibo_content)
            # print('tags:', len(tags))
            for i in range(len(tags)):
                tag_text = tags[i].text
                # print(tag_text)
                weibo_content = weibo_content.replace(tag_text, '',1)

            print("weibo content：",weibo_content)
            print("number of photos：",len(pics))
            
            value = [weibo_content]
            write_excel_xls_append_norepeat(path, value)
            img_download(pics,image_path,path)

#爬取数据
def get_data(path,userId):
    basic_url = 'https://weibo.com/u/'+userId+'?is_search=0&visible=0&is_all=1&is_tag=0&profile_ftype=1'
    first_url = basic_url + '&page=1'
    # print(first_url)
    driver.get(first_url)
    before = 0 
    after = 0
    N = 0
    #找到微博页数，只有拉到浏览器最下面才知道一共有多少页的微博
    while True:
        before = after
        Transfer_Clicks(driver)#下滚操作，加载更多内容
        time.sleep(1)
        es = driver.find_elements_by_xpath("//div[@action-data='cur_visible=0' and @action-type='feed_list_item']")
        print("The total number of weibo in this page ：%d, N is now：%d, there is no new weibo when N reaches 3" % (len(es),N))
        after = len(es)
        if after > before:
            N = 0
        if after == before:        
            N = N + 1
        if N == 5:
            print("The total number of weibo in this page ：%d" % after)
            break
    time.sleep(1)

    #获取页数
    pages = driver.find_elements_by_xpath("//div[@class='W_pages']//li")
    pages = len(pages)
    print("there are ",pages,"pages")
    
    #当只有1页时，只用输出当页的内容即可
    if pages is 0:
        print('only one page')
        url = basic_url + '&page='+ str(1)+'#feedtop'

        driver.get(url)
        before = 0 
        after = 0
        N = 0
        while True:
            before = after
            Transfer_Clicks(driver)
            time.sleep(1)
            es = driver.find_elements_by_xpath("//div[@action-data='cur_visible=0' and @action-type='feed_list_item']")
            print("The total number of weibo in this page ：%d, N is now：%d, there is no new weibo when N reaches 3" % (len(es),N))
            after = len(es)
            if after > before:
                N = 0
            if after == before:        
                N = N + 1
            if N == 5:
                break
        elems = driver.find_elements_by_xpath("//div[@action-data='cur_visible=0' and @action-type='feed_list_item']")
        insert_data(path,elems,userId)
        time.sleep(1)
    #如果不只一页，就要访问每一页来获取数据
    for page in range(pages):
        print('**********************************')
        url = basic_url + '&page='+ str(page+1)+'#feedtop'

        driver.get(url)
        before = 0 
        after = 0
        N = 0
        #同样，在每一页微博都要向下滚直到加载出当页的所有内容
        while True:
            before = after
            Transfer_Clicks(driver)
            time.sleep(1)
            es = driver.find_elements_by_xpath("//div[@action-data='cur_visible=0' and @action-type='feed_list_item']")
            print("The total number of weibo in this page ：%d, N is now：%d, there is no new weibo when N reaches 3" % (len(es),N))
            after = len(es)
            if after > before:
                N = 0
            if after == before:        
                N = N + 1
            if N == 5:
                print("This is page number %d ，there are %d pages in total" % (page+1,pages))
                break
        elems = driver.find_elements_by_xpath("//div[@action-data='cur_visible=0' and @action-type='feed_list_item']")
        insert_data(path,elems,userId)
        time.sleep(15)



from selenium import webdriver
import time

#全局变量
driver = webdriver.Chrome()
driver.maximize_window()

#基本信息
def visitUserInfo(userId):
    driver.get('http://weibo.com/u/' + userId)

    print('********************')   
    print('show user information：')
    
    time.sleep(5)
    nickname = driver.find_element_by_xpath("//h1[@class='username']").text
    print('nickname:' + nickname)


if __name__ == '__main__':
    username = '你的微博账号'             # 输入微博账号
    password = '账号密码'             # 输入密码


    loginWeibo(username, password)     
    time.sleep(15)

    uid = '你想爬的用户id'
    visitUserInfo(uid)

    path = "depression/csdn/"+uid+".xlsx"
    
    #创建文件
    if os.path.exists(path):
        print("The file has already been created.")
    else:
        print("Create a new file.")
        write_excel_xls(path)
    
    get_data(path,uid) 